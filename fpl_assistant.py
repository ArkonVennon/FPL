"""
FPL Analytics Assistant
=======================
A machine learning-powered Fantasy Premier League assistant that combines:
  - Live FPL API data ingestion
  - Position-specific player performance analysis
  - Custom fixture difficulty rating (SUPER_EASY → SUPER_HARD)
  - GradientBoosting ML model for transfer recommendations
  - Optional LLM sentiment layer via OpenAI API

Usage:
    python fpl_assistant.py

Set your TEAM_ID in config.py (or directly below) and optionally supply
an OPENAI_API_KEY environment variable to unlock conversational analysis.
"""

import os
import warnings

import numpy as np
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Optional: OpenAI integration (install: pip install openai)
# ---------------------------------------------------------------------------
try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False


# ===========================================================================
# Data Ingestion
# ===========================================================================

class FPLDataIngestion:
    """Fetch data exclusively from the official FPL REST API."""

    BASE_URL = "https://fantasy.premierleague.com/api/"

    def __init__(self):
        self.session = requests.Session()

    def _get(self, url: str):
        """GET helper with basic error handling."""
        try:
            response = self.session.get(url, timeout=10)
            response.raise_for_status()
            return response.json()
        except Exception as exc:
            print(f"[ERROR] Request failed for {url}: {exc}")
            return None

    def get_fpl_team_data(self, team_id: int, gameweek: int | None = None):
        """Return picks for a specific GW, or full entry history."""
        if gameweek:
            url = f"{self.BASE_URL}entry/{team_id}/event/{gameweek}/picks/"
        else:
            url = f"{self.BASE_URL}entry/{team_id}/history/"
        return self._get(url)

    def get_all_gameweek_data(self, team_id: int, max_gameweek: int = 38) -> pd.DataFrame:
        """Download every completed gameweek for the given team."""
        records = []
        print(f"Fetching gameweek history for team {team_id} …")

        for gw in range(1, max_gameweek + 1):
            data = self.get_fpl_team_data(team_id, gw)
            if not data:
                break  # No more completed GWs
            history = data.get("entry_history", {})
            records.append(
                {
                    "gameweek": gw,
                    "points": history.get("points", 0),
                    "total_points": history.get("total_points", 0),
                    "rank": history.get("rank", 0),
                    "bank": history.get("bank", 0),
                    "value": history.get("value", 0),
                    "event_transfers": history.get("event_transfers", 0),
                    "event_transfers_cost": history.get("event_transfers_cost", 0),
                    "picks": data.get("picks", []),
                }
            )
            print(f"  ✓ GW{gw}: {history.get('points', 0)} pts")

        return pd.DataFrame(records)

    def get_player_details(self) -> tuple[pd.DataFrame | None, pd.DataFrame | None]:
        """Return (players_df, teams_df) from the bootstrap-static endpoint."""
        data = self._get(f"{self.BASE_URL}bootstrap-static/")
        if not data:
            return None, None

        players_df = pd.DataFrame(data["elements"])
        teams_df = pd.DataFrame(data["teams"])

        for col in ["expected_goals", "expected_assists",
                    "expected_goal_involvements", "expected_goals_conceded"]:
            players_df[col] = pd.to_numeric(players_df[col], errors="coerce")

        return players_df, teams_df

    def get_player_gameweek_history(self, player_id: int) -> pd.DataFrame | None:
        """Return per-GW history for a single player."""
        data = self._get(f"{self.BASE_URL}element-summary/{player_id}/")
        if not data:
            return None
        return pd.DataFrame(data["history"])


# ===========================================================================
# Fixture Difficulty Rating
# ===========================================================================

class FixtureDifficultyAnalyzer:
    """
    Custom FDR that replaces the FPL default.

    Difficulty scale
    ----------------
    SUPER_EASY  (1) — mid/bottom-tier team vs. consistently weak opponent
    EASY        (2) — same tier or bottom vs. non-top
    MODERATE    (3) — mixed-tier matchup
    HARD        (4) — non-top team facing a top-5 side
    SUPER_HARD  (5) — non-top team facing a team that has been in top-5 for ≥5 GWs
    """

    BASE_URL = "https://fantasy.premierleague.com/api/"

    DIFFICULTY_SCORES = {
        "SUPER_EASY": 1,
        "EASY": 2,
        "MODERATE": 3,
        "HARD": 4,
        "SUPER_HARD": 5,
    }

    def __init__(self):
        self.session = requests.Session()
        self.league_table: pd.DataFrame | None = None
        self.fixtures: pd.DataFrame | None = None
        self.teams_df: pd.DataFrame | None = None

    def _get(self, url: str):
        try:
            r = self.session.get(url, timeout=10)
            r.raise_for_status()
            return r.json()
        except Exception as exc:
            print(f"[ERROR] {url}: {exc}")
            return None

    def fetch_league_table(self) -> pd.DataFrame | None:
        """Populate self.league_table from bootstrap-static."""
        data = self._get(f"{self.BASE_URL}bootstrap-static/")
        if not data:
            return None
        df = pd.DataFrame(data["teams"]).sort_values("position")
        self.league_table = df[["id", "name", "short_name", "position", "strength"]]
        self.teams_df = self.league_table
        print(f"  ✓ League table: {len(self.league_table)} teams")
        return self.league_table

    def fetch_fixtures(self) -> pd.DataFrame | None:
        """Populate self.fixtures."""
        data = self._get(f"{self.BASE_URL}fixtures/")
        if not data:
            return None
        self.fixtures = pd.DataFrame(data)
        print(f"  ✓ Fixtures: {len(self.fixtures)} rows")
        return self.fixtures

    # ------------------------------------------------------------------
    # Tier / difficulty helpers
    # ------------------------------------------------------------------

    def _tier(self, position: int) -> str:
        if position <= 5:
            return "TOP"
        if position <= 15:
            return "MIDDLE"
        return "BOTTOM"

    def _base_difficulty(self, team_pos: int, opp_pos: int) -> str:
        team_tier = self._tier(team_pos)
        opp_tier = self._tier(opp_pos)
        if team_tier == opp_tier:
            return "EASY"
        if opp_tier == "TOP" and team_tier != "TOP":
            return "HARD"
        if opp_tier == "BOTTOM" and team_tier != "BOTTOM":
            return "EASY"
        return "MODERATE"

    def _opponent_form(self, positions: list[int]) -> str:
        if not positions:
            return "NORMAL"
        if all(p <= 5 for p in positions):
            return "SUPER_TOP"
        if all(p >= 16 for p in positions):
            return "SUPER_BOTTOM"
        return "NORMAL"

    def _escalate(self, base: str, form: str) -> str:
        if base == "HARD" and form == "SUPER_TOP":
            return "SUPER_HARD"
        if base == "EASY" and form == "SUPER_BOTTOM":
            return "SUPER_EASY"
        return base

    def get_fixture_difficulty(self, team_id: int, opp_id: int, current_gw: int) -> dict | None:
        """Return full difficulty breakdown for one fixture."""
        if self.league_table is None:
            return None
        t = self.league_table[self.league_table["id"] == team_id]
        o = self.league_table[self.league_table["id"] == opp_id]
        if t.empty or o.empty:
            return None

        team_pos = int(t["position"].values[0])
        opp_pos = int(o["position"].values[0])
        opp_last5 = [opp_pos] * min(5, current_gw)  # simplified; track history for full impl.

        base = self._base_difficulty(team_pos, opp_pos)
        form = self._opponent_form(opp_last5)
        final = self._escalate(base, form)

        return {
            "team_position": team_pos,
            "opponent_position": opp_pos,
            "team_tier": self._tier(team_pos),
            "opponent_tier": self._tier(opp_pos),
            "base_difficulty": base,
            "opponent_form": form,
            "final_difficulty": final,
        }

    def _get_fixtures_for_team(self, team_id: int, finished: bool, n: int) -> pd.DataFrame:
        if self.fixtures is None:
            return pd.DataFrame()
        mask = (
            (self.fixtures["finished"] == finished)
            & ((self.fixtures["team_h"] == team_id) | (self.fixtures["team_a"] == team_id))
        )
        return (
            self.fixtures[mask]
            .sort_values("event", ascending=not finished)
            .head(n)
        )

    def _team_name(self, team_id: int) -> str:
        if self.teams_df is None:
            return f"Team {team_id}"
        row = self.teams_df[self.teams_df["id"] == team_id]
        return row["short_name"].values[0] if not row.empty else f"Team {team_id}"

    def analyze_player_fixture_difficulty(self, player_row: pd.Series, current_gw: int) -> dict:
        """Compute past / upcoming fixture difficulty for a player."""
        team_id = player_row["team"]

        def _analyze(fixtures_df: pd.DataFrame) -> list[dict]:
            result = []
            for _, fx in fixtures_df.iterrows():
                opp_id = fx["team_a"] if fx["team_h"] == team_id else fx["team_h"]
                diff = self.get_fixture_difficulty(team_id, opp_id, current_gw)
                if diff:
                    result.append(
                        {
                            "gameweek": fx["event"],
                            "opponent_id": opp_id,
                            "opponent_name": self._team_name(opp_id),
                            "is_home": fx["team_h"] == team_id,
                            "difficulty": diff["final_difficulty"],
                            "opponent_position": diff["opponent_position"],
                        }
                    )
            return result

        past = _analyze(self._get_fixtures_for_team(team_id, finished=True, n=5))
        upcoming = _analyze(self._get_fixtures_for_team(team_id, finished=False, n=5))

        def _avg(items):
            scores = [self.DIFFICULTY_SCORES.get(i["difficulty"], 3) for i in items]
            return float(np.mean(scores)) if scores else 3.0

        past_avg = _avg(past)
        upcoming_avg = _avg(upcoming)

        return {
            "player_name": player_row["web_name"],
            "team_id": team_id,
            "past_fixtures": past,
            "upcoming_fixtures": upcoming,
            "past_avg_difficulty": past_avg,
            "upcoming_avg_difficulty": upcoming_avg,
            "fixture_swing": upcoming_avg - past_avg,
        }

    def sentiment_text(self, analysis: dict) -> str:
        """Human-readable summary of fixture swing."""
        p, u, s = analysis["past_avg_difficulty"], analysis["upcoming_avg_difficulty"], analysis["fixture_swing"]
        past_label = "Easy recent" if p <= 2 else ("Tough recent" if p >= 4 else "Moderate recent")
        up_label = "excellent upcoming" if u <= 2 else ("difficult upcoming" if u >= 4 else "moderate upcoming")
        swing_label = "IMPROVING ⬆️" if s <= -1 else ("WORSENING ⬇️" if s >= 1 else "stable")
        return f"{past_label} | {up_label} | {swing_label}"


# ===========================================================================
# Player Performance Analyzers
# ===========================================================================

def _to_num(val, default=0):
    return pd.to_numeric(val, errors="coerce") if val is not None else default


class FPLTrendAnalyzer:
    """Position-specific performance analysis using FPL API stats."""

    def _goalkeeper_stats(self, row) -> dict:
        mins = _to_num(row.get("minutes", 0))
        saves = _to_num(row.get("saves", 0))
        gc = _to_num(row.get("goals_conceded", 0))
        xgc = _to_num(row.get("expected_goals_conceded", 0))
        starts = _to_num(row.get("starts", 0))
        cs = _to_num(row.get("clean_sheets", 0))
        n90 = mins / 90.0 if mins > 0 else 0
        return {
            "clean_sheets": cs,
            "goals_conceded": gc,
            "saves": saves,
            "minutes": mins,
            "starts": starts,
            "form": _to_num(row.get("form", 0)),
            "points_per_game": _to_num(row.get("points_per_game", 0)),
            "expected_goals_conceded": xgc,
            "bonus": _to_num(row.get("bonus", 0)),
            "penalties_saved": _to_num(row.get("penalties_saved", 0)),
            "saves_per_90": saves / n90 if n90 else 0,
            "goals_conceded_per_90": gc / n90 if n90 else 0,
            "clean_sheet_rate": (cs / starts * 100) if starts else 0,
            "gc_vs_xgc": gc - xgc if xgc else 0,
        }

    def _defender_stats(self, row) -> dict:
        mins = _to_num(row.get("minutes", 0))
        goals = _to_num(row.get("goals_scored", 0))
        assists = _to_num(row.get("assists", 0))
        gc = _to_num(row.get("goals_conceded", 0))
        xgc = _to_num(row.get("expected_goals_conceded", 0))
        n90 = mins / 90.0 if mins > 0 else 0
        return {
            "clean_sheets": _to_num(row.get("clean_sheets", 0)),
            "goals_conceded": gc,
            "expected_goals_conceded": xgc,
            "goals_scored": goals,
            "assists": assists,
            "expected_goals": _to_num(row.get("expected_goals", 0)),
            "expected_assists": _to_num(row.get("expected_assists", 0)),
            "minutes": mins,
            "starts": _to_num(row.get("starts", 0)),
            "form": _to_num(row.get("form", 0)),
            "points_per_game": _to_num(row.get("points_per_game", 0)),
            "bonus": _to_num(row.get("bonus", 0)),
            "bps": _to_num(row.get("bps", 0)),
            "influence": _to_num(row.get("influence", 0)),
            "yellow_cards": _to_num(row.get("yellow_cards", 0)),
            "red_cards": _to_num(row.get("red_cards", 0)),
            "attacking_returns_per_90": (goals + assists) / n90 if n90 else 0,
            "gc_vs_xgc": gc - xgc if xgc else 0,
        }

    def _midfielder_stats(self, row) -> dict:
        mins = _to_num(row.get("minutes", 0))
        goals = _to_num(row.get("goals_scored", 0))
        assists = _to_num(row.get("assists", 0))
        xgi = _to_num(row.get("expected_goal_involvements", 0))
        n90 = mins / 90.0 if mins > 0 else 0
        actual_gi = goals + assists
        overperf = actual_gi - xgi if xgi else 0
        return {
            "goals_scored": goals,
            "assists": assists,
            "expected_goals": _to_num(row.get("expected_goals", 0)),
            "expected_assists": _to_num(row.get("expected_assists", 0)),
            "expected_goal_involvements": xgi,
            "creativity": _to_num(row.get("creativity", 0)),
            "threat": _to_num(row.get("threat", 0)),
            "ict_index": _to_num(row.get("ict_index", 0)),
            "influence": _to_num(row.get("influence", 0)),
            "minutes": mins,
            "starts": _to_num(row.get("starts", 0)),
            "form": _to_num(row.get("form", 0)),
            "points_per_game": _to_num(row.get("points_per_game", 0)),
            "bonus": _to_num(row.get("bonus", 0)),
            "bps": _to_num(row.get("bps", 0)),
            "yellow_cards": _to_num(row.get("yellow_cards", 0)),
            "red_cards": _to_num(row.get("red_cards", 0)),
            "goal_involvements_per_90": actual_gi / n90 if n90 else 0,
            "xgi_overperformance": overperf,
            "xgi_overperformance_pct": (overperf / xgi * 100) if xgi else 0,
        }

    def _forward_stats(self, row) -> dict:
        mins = _to_num(row.get("minutes", 0))
        goals = _to_num(row.get("goals_scored", 0))
        xg = _to_num(row.get("expected_goals", 0))
        n90 = mins / 90.0 if mins > 0 else 0
        return {
            "goals_scored": goals,
            "assists": _to_num(row.get("assists", 0)),
            "expected_goals": xg,
            "expected_assists": _to_num(row.get("expected_assists", 0)),
            "expected_goal_involvements": _to_num(row.get("expected_goal_involvements", 0)),
            "threat": _to_num(row.get("threat", 0)),
            "ict_index": _to_num(row.get("ict_index", 0)),
            "influence": _to_num(row.get("influence", 0)),
            "creativity": _to_num(row.get("creativity", 0)),
            "minutes": mins,
            "starts": _to_num(row.get("starts", 0)),
            "form": _to_num(row.get("form", 0)),
            "points_per_game": _to_num(row.get("points_per_game", 0)),
            "bonus": _to_num(row.get("bonus", 0)),
            "bps": _to_num(row.get("bps", 0)),
            "yellow_cards": _to_num(row.get("yellow_cards", 0)),
            "red_cards": _to_num(row.get("red_cards", 0)),
            "goals_per_90": goals / n90 if n90 else 0,
            "xg_conversion_rate": (goals / xg * 100) if xg else 0,
        }

    def analyze_player(self, player_row: pd.Series) -> dict:
        """Return position-specific performance dict for one player."""
        pos = player_row.get("element_type")
        dispatch = {1: self._goalkeeper_stats, 2: self._defender_stats,
                    3: self._midfielder_stats, 4: self._forward_stats}
        pos_stats = dispatch.get(pos, lambda _: {})(player_row)

        base = {
            "player_id": player_row.get("id"),
            "name": player_row.get("web_name"),
            "position": pos,
            "team": player_row.get("team"),
            "total_points": _to_num(player_row.get("total_points", 0)),
            "now_cost": _to_num(player_row.get("now_cost", 0)) / 10,
        }
        base.update(pos_stats)
        return base

    def analyze_squad(self, picks: list[dict], players_df: pd.DataFrame) -> pd.DataFrame:
        """Analyse every player in the current squad."""
        rows = []
        for pick in picks:
            pid = pick.get("element")
            row = players_df[players_df["id"] == pid]
            if row.empty:
                continue
            rows.append(self.analyze_player(row.iloc[0]))
        return pd.DataFrame(rows)

    def identify_transfer_candidates(self, df: pd.DataFrame) -> pd.DataFrame:
        """Flag underperforming players with a severity score."""
        candidates = []
        for _, p in df.iterrows():
            issues, score = [], 0
            pos = p.get("position")

            if pos == 1:
                if p.get("saves_per_90", 99) < 2.0:
                    issues.append(f"Low saves/90: {p.get('saves_per_90', 0):.2f}")
                    score += 2
                if p.get("clean_sheet_rate", 100) < 20:
                    issues.append(f"Low CS rate: {p.get('clean_sheet_rate', 0):.1f}%")
                    score += 2
            elif pos == 2:
                if p.get("gc_vs_xgc", 0) > 2:
                    issues.append(f"Conceding > xGC: +{p.get('gc_vs_xgc', 0):.2f}")
                    score += 2
            elif pos == 3:
                if p.get("xgi_overperformance_pct", 0) < -20:
                    issues.append(f"xGI underperformance: {p.get('xgi_overperformance_pct', 0):.1f}%")
                    score += 2
            elif pos == 4:
                if p.get("xg_conversion_rate", 100) < 60:
                    issues.append(f"Low xG conversion: {p.get('xg_conversion_rate', 0):.1f}%")
                    score += 2

            if p.get("minutes", 9999) < 500:
                issues.append(f"Low minutes: {int(p.get('minutes', 0))}")
                score += 3
            if p.get("points_per_game", 99) < 2.5:
                issues.append(f"Low PPG: {p.get('points_per_game', 0):.1f}")
                score += 2
            if p.get("form", 99) < 2.0:
                issues.append(f"Poor form: {p.get('form', 0):.1f}")
                score += 1
            if p.get("fixture_swing", 0) > 1.5:
                issues.append(f"Fixtures worsening: +{p.get('fixture_swing', 0):.1f}")
                score += 2

            if issues:
                candidates.append({
                    "name": p["name"],
                    "position": pos,
                    "severity": score,
                    "issues": " | ".join(issues),
                    "total_points": p.get("total_points", 0),
                })

        if candidates:
            return pd.DataFrame(candidates).sort_values("severity", ascending=False)
        return pd.DataFrame()


class FPLTrendAnalyzerEnhanced(FPLTrendAnalyzer):
    """Extends FPLTrendAnalyzer with optional fixture difficulty overlay."""

    def __init__(self, fixture_analyzer: FixtureDifficultyAnalyzer | None = None):
        self.fixture_analyzer = fixture_analyzer

    def analyze_player_with_fixtures(self, player_row: pd.Series, current_gw: int) -> dict:
        data = self.analyze_player(player_row)
        if not self.fixture_analyzer:
            return data
        try:
            fa = self.fixture_analyzer.analyze_player_fixture_difficulty(player_row, current_gw)
            data.update({
                "past_avg_difficulty": fa["past_avg_difficulty"],
                "upcoming_avg_difficulty": fa["upcoming_avg_difficulty"],
                "fixture_swing": fa["fixture_swing"],
                "fixture_sentiment": self.fixture_analyzer.sentiment_text(fa),
                "upcoming_fixtures_detail": fa["upcoming_fixtures"],
                "past_fixtures_detail": fa["past_fixtures"],
            })
        except Exception as exc:
            print(f"  [WARN] Fixture analysis failed for {player_row['web_name']}: {exc}")
            data.update({"past_avg_difficulty": 3, "upcoming_avg_difficulty": 3,
                         "fixture_swing": 0, "fixture_sentiment": "N/A"})
        return data

    def analyze_squad_with_fixtures(self, picks: list[dict],
                                     players_df: pd.DataFrame, current_gw: int) -> pd.DataFrame:
        rows = []
        for pick in picks:
            pid = pick.get("element")
            row = players_df[players_df["id"] == pid]
            if row.empty:
                continue
            rows.append(self.analyze_player_with_fixtures(row.iloc[0], current_gw))
        return pd.DataFrame(rows)


# ===========================================================================
# Excel Export
# ===========================================================================

class ExcelExporter:
    """Write analysis results to a formatted .xlsx file."""

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    @classmethod
    def export(cls, performance_df: pd.DataFrame, transfer_df: pd.DataFrame,
               output_path: str) -> None:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb = Workbook()
        ws_perf = wb.active
        ws_perf.title = "Player Performance"
        ws_transfer = wb.create_sheet("Transfer Targets")

        cls._write_performance_sheet(ws_perf, performance_df)
        cls._write_transfer_sheet(ws_transfer, transfer_df)
        wb.save(output_path)
        print(f"  ✓ Excel saved → {output_path}")

    @classmethod
    def _auto_col_width(cls, ws) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col:
                if isinstance(cell, MergedCell):
                    continue
                col_letter = col_letter or cell.column_letter
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_len + 2, 45)

    @classmethod
    def _write_performance_sheet(cls, ws, df: pd.DataFrame) -> None:
        ws["A1"] = "Player Performance Analysis (FPL Stats + Fixtures)"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:H1")

        if df.empty:
            return

        pos_map = {1: "GK", 2: "DEF", 3: "MID", 4: "FWD"}
        df = df.copy()
        df["position_name"] = df["position"].map(pos_map)

        wanted = ["name", "position_name", "total_points", "form", "points_per_game",
                  "minutes", "past_avg_difficulty", "upcoming_avg_difficulty",
                  "fixture_swing", "fixture_sentiment"]
        display_df = df[[c for c in wanted if c in df.columns]]

        for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), start=3):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = cls.THIN_BORDER
                if r_idx == 3:
                    cell.font = cls.HEADER_FONT
                    cell.fill = cls.HEADER_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif isinstance(val, (int, float)):
                    cell.number_format = "0.00"
        cls._auto_col_width(ws)

    @classmethod
    def _write_transfer_sheet(cls, ws, df: pd.DataFrame) -> None:
        ws["A1"] = "Transfer Targets — Players to Consider Removing"
        ws["A1"].font = Font(bold=True, size=14, color="FF0000")
        ws.merge_cells("A1:E1")

        if df.empty:
            ws["A3"] = "No transfer targets identified — all players performing well!"
            ws["A3"].font = Font(color="00B050", bold=True)
            return

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = cls.THIN_BORDER
                if r_idx == 3:
                    cell.font = cls.HEADER_FONT
                    cell.fill = cls.HEADER_FILL
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif c_idx == 3 and isinstance(val, (int, float)):
                    color = "FFC7CE" if val >= 5 else ("FFEB9C" if val >= 3 else None)
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                if isinstance(val, (int, float)) and c_idx != 2:
                    cell.number_format = "0.00"
        cls._auto_col_width(ws)


# ===========================================================================
# Optional: OpenAI / ChatGPT Integration
# ===========================================================================

class ChatGPTIntegration:
    """
    Wraps OpenAI chat completions for two purposes:

    1. Conversational Q&A about your squad (start_chat mode).
    2. Per-player sentiment scoring that adjusts ML recommendations.

    Requires the ``openai`` package and a valid OPENAI_API_KEY env var.
    """

    SENTIMENT_WEIGHTS = {
        "very_positive": 1.3,
        "positive": 1.15,
        "neutral": 1.0,
        "negative": 0.85,
        "very_negative": 0.7,
    }

    def __init__(self, api_key: str):
        if not OPENAI_AVAILABLE:
            raise ImportError("openai package not installed. Run: pip install openai")
        openai.api_key = api_key
        self.history: list[dict] = []

    def build_context(self, analysis_df: pd.DataFrame, transfer_df: pd.DataFrame,
                      recs_text: str, players_df: pd.DataFrame | None = None) -> str:
        ctx = (
            "You are an expert FPL analyst. You have just completed a detailed analysis.\n\n"
            f"TEAM ANALYSIS:\n{analysis_df.to_string(index=False) if not analysis_df.empty else 'N/A'}\n\n"
            f"TRANSFER TARGETS:\n{transfer_df.to_string(index=False) if not transfer_df.empty else 'None'}\n\n"
            f"RECOMMENDATIONS:\n{recs_text}\n"
        )
        if players_df is not None:
            ctx += (
                f"\nYou have access to stats for all {len(players_df)} FPL players. "
                "When asked about any player, use their actual data."
            )
        return ctx

    def chat(self, user_msg: str, context: str | None = None,
             players_df: pd.DataFrame | None = None) -> str:
        """Send a message and return the assistant reply."""
        enhanced = user_msg
        if players_df is not None:
            for _, row in players_df.iterrows():
                if str(row["web_name"]).lower() in user_msg.lower():
                    pos = ["GK", "DEF", "MID", "FWD"][int(row["element_type"]) - 1]
                    enhanced += (
                        f"\n\n[Stats: {row['web_name']} | {pos} | £{float(row['now_cost'])/10}m | "
                        f"Pts:{row['total_points']} PPG:{row['points_per_game']} Form:{row['form']} "
                        f"Mins:{row['minutes']} xG:{row['expected_goals']} xA:{row['expected_assists']}]"
                    )
                    break

        self.history.append({"role": "user", "content": enhanced})
        messages = []
        if context:
            messages.append({"role": "system", "content": context})
        messages.extend(self.history)

        try:
            resp = openai.chat.completions.create(
                model="gpt-4o-mini",
                messages=messages,
                temperature=0.7,
                max_tokens=500,
            )
            reply = resp.choices[0].message.content
            self.history.append({"role": "assistant", "content": reply})
            return reply
        except Exception as exc:
            return f"[ChatGPT error] {exc}"

    def get_player_sentiment(self, player_data: dict,
                              fixture_data: dict | None = None) -> dict:
        """Return sentiment label + weight multiplier for a player."""
        pos = player_data.get("position", 1)
        pos_name = ["GK", "DEF", "MID", "FWD"][pos - 1]

        prompt = (
            f"Analyse this FPL {pos_name} and return:\n"
            f"SENTIMENT: [very_positive/positive/neutral/negative/very_negative]\n"
            f"REASONING: [one sentence]\n\n"
            f"Name: {player_data.get('name')} | Cost: £{player_data.get('cost', 0)}m | "
            f"Pts: {player_data.get('total_points')} | PPG: {player_data.get('points_per_game', 0):.1f} | "
            f"Form: {player_data.get('form', 0):.1f} | Mins: {player_data.get('minutes')}\n"
            f"xG: {player_data.get('expected_goals', 0):.2f} | "
            f"xA: {player_data.get('expected_assists', 0):.2f} | "
            f"xGI: {player_data.get('expected_goal_involvements', 0):.2f}\n"
        )
        if fixture_data:
            prompt += (
                f"Upcoming fixture avg difficulty: {fixture_data.get('upcoming_avg_difficulty', 3):.1f}/5 | "
                f"Swing: {fixture_data.get('fixture_swing', 0):+.1f}\n"
            )

        try:
            resp = openai.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are a concise FPL expert."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.4,
                max_tokens=80,
            )
            text = resp.choices[0].message.content
            sentiment = next(
                (k for k in self.SENTIMENT_WEIGHTS if k in text.lower()), "neutral"
            )
            return {"sentiment": sentiment, "weight": self.SENTIMENT_WEIGHTS[sentiment],
                    "reasoning": text}
        except Exception as exc:
            return {"sentiment": "neutral", "weight": 1.0, "reasoning": str(exc)}

    def reset(self):
        self.history = []


# ===========================================================================
# ML Recommendation Engine
# ===========================================================================

FEATURE_COLS = [
    "form", "points_per_game", "total_points",
    "expected_goals", "expected_assists", "expected_goal_involvements",
    "expected_goals_conceded", "goals_scored", "assists", "clean_sheets",
    "bonus", "bps", "influence", "creativity", "threat", "ict_index",
    "minutes", "now_cost", "starts", "saves",
]


class FPLRecommendationEngine:
    """
    GradientBoosting model that predicts total_points, then optionally
    adjusts the score with an LLM sentiment multiplier.
    """

    def __init__(self, chatgpt: ChatGPTIntegration | None = None,
                 fixture_analyzer: FixtureDifficultyAnalyzer | None = None):
        self.model = GradientBoostingRegressor(n_estimators=150, learning_rate=0.1,
                                               max_depth=4, random_state=42)
        self.scaler = StandardScaler()
        self.chatgpt = chatgpt
        self.fixture_analyzer = fixture_analyzer
        self._feature_names: list[str] = []

    def _prepare_features(self, df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
        df = df.copy()
        existing = [f for f in FEATURE_COLS if f in df.columns]
        for col in existing:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df[existing], existing

    def train(self, players_df: pd.DataFrame) -> None:
        X, names = self._prepare_features(players_df)
        y = pd.to_numeric(players_df["total_points"], errors="coerce").fillna(0)
        self._feature_names = names
        self.model.fit(self.scaler.fit_transform(X), y)
        print(f"  ✓ Model trained on {len(names)} features, {len(X)} players")

    def recommend(self, players_df: pd.DataFrame, exclude_ids: list[int],
                  position: int, budget: float,
                  n: int = 3, current_gw: int = 24) -> pd.DataFrame:
        """Return top-n recommendations for a position within budget."""
        pool = players_df[
            (~players_df["id"].isin(exclude_ids))
            & (players_df["element_type"] == position)
            & (pd.to_numeric(players_df["now_cost"], errors="coerce") <= budget * 10)
            & (pd.to_numeric(players_df["minutes"], errors="coerce") > 200)
        ].copy()

        if pool.empty:
            print(f"  No candidates for position {position} within £{budget}m")
            return pd.DataFrame()

        X, _ = self._prepare_features(pool)
        pool["ml_score"] = self.model.predict(self.scaler.transform(X))
        pool = pool.nlargest(n, "ml_score")

        records = []
        for _, p in pool.iterrows():
            fixture_data = None
            if self.fixture_analyzer:
                try:
                    fixture_data = self.fixture_analyzer.analyze_player_fixture_difficulty(
                        p, current_gw)
                except Exception:
                    pass

            if self.chatgpt:
                player_data = {
                    "name": p["web_name"], "position": position,
                    "cost": _to_num(p["now_cost"]) / 10,
                    "total_points": _to_num(p["total_points"]),
                    "points_per_game": _to_num(p.get("points_per_game", 0)),
                    "form": _to_num(p.get("form", 0)),
                    "minutes": _to_num(p.get("minutes", 0)),
                    "expected_goals": _to_num(p.get("expected_goals", 0)),
                    "expected_assists": _to_num(p.get("expected_assists", 0)),
                    "expected_goal_involvements": _to_num(p.get("expected_goal_involvements", 0)),
                }
                sentiment = self.chatgpt.get_player_sentiment(player_data, fixture_data)
            else:
                sentiment = {"sentiment": "neutral", "weight": 1.0, "reasoning": "—"}

            rec = {
                "name": p["web_name"],
                "cost": _to_num(p["now_cost"]) / 10,
                "total_points": _to_num(p["total_points"]),
                "form": _to_num(p.get("form", 0)),
                "ml_score": p["ml_score"],
                "sentiment": sentiment["sentiment"],
                "final_score": p["ml_score"] * sentiment["weight"],
            }
            if fixture_data:
                rec["upcoming_difficulty"] = fixture_data.get("upcoming_avg_difficulty", 3)
                rec["fixture_swing"] = fixture_data.get("fixture_swing", 0)
            records.append(rec)

        return pd.DataFrame(records).sort_values("final_score", ascending=False)


# ===========================================================================
# Orchestrator
# ===========================================================================

class FPLAssistant:
    """
    Top-level orchestrator.  Typical workflow::

        assistant = FPLAssistant(team_id=YOUR_ID)
        assistant.initialize()
        assistant.run_analysis()
        assistant.export_data()
        # Optional: assistant.start_chat()
    """

    POSITION_NAMES = {1: "GOALKEEPER", 2: "DEFENDER", 3: "MIDFIELDER", 4: "FORWARD"}

    def __init__(self, team_id: int, openai_api_key: str | None = None, output_dir: str = "./output"):
        self.team_id = team_id
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

        self.ingestion = FPLDataIngestion()
        self.fixture_analyzer = FixtureDifficultyAnalyzer()
        self.analyzer = FPLTrendAnalyzerEnhanced(fixture_analyzer=self.fixture_analyzer)

        self.chatgpt: ChatGPTIntegration | None = None
        if openai_api_key and OPENAI_AVAILABLE:
            self.chatgpt = ChatGPTIntegration(openai_api_key)
            print("  ✓ OpenAI integration enabled")
        elif openai_api_key and not OPENAI_AVAILABLE:
            print("  [WARN] openai package not installed — chat features disabled")

        self.recommender = FPLRecommendationEngine(
            chatgpt=self.chatgpt,
            fixture_analyzer=self.fixture_analyzer,
        )

        self.team_data: pd.DataFrame | None = None
        self.players_df: pd.DataFrame | None = None
        self.teams_df: pd.DataFrame | None = None
        self.comparison: pd.DataFrame = pd.DataFrame()
        self.transfer_targets: pd.DataFrame = pd.DataFrame()
        self.all_recommendations: dict = {}

    def initialize(self) -> None:
        print("\n=== FPL Assistant initialising ===")
        print("1. Fetching player data …")
        self.players_df, self.teams_df = self.ingestion.get_player_details()
        if self.players_df is not None:
            print(f"  ✓ {len(self.players_df)} players loaded")

        print("2. Fetching gameweek history …")
        self.team_data = self.ingestion.get_all_gameweek_data(self.team_id)
        if self.team_data is not None:
            print(f"  ✓ {len(self.team_data)} gameweeks loaded")

        print("3. Fetching league table and fixtures …")
        self.fixture_analyzer.fetch_league_table()
        self.fixture_analyzer.fetch_fixtures()
        print("=== Initialisation complete ===\n")

    def run_analysis(self) -> dict:
        if self.team_data is None or self.players_df is None:
            raise RuntimeError("Call initialize() before run_analysis().")

        current_gw = self.team_data.iloc[-1]
        picks = current_gw["picks"]
        current_ids = [p["element"] for p in picks]

        print(f"\n{'='*70}")
        print(f"GW{current_gw['gameweek']} — {current_gw['points']} pts "
              f"(total {current_gw['total_points']}) | "
              f"Value £{current_gw['value']/10}m | Bank £{current_gw['bank']/10}m")
        print(f"{'='*70}\n")

        # --- Squad analysis ---
        print("Analysing squad …")
        self.comparison = self.analyzer.analyze_squad_with_fixtures(
            picks, self.players_df, current_gw["gameweek"]
        )
        show_cols = [c for c in ["name", "position", "total_points", "form",
                                  "points_per_game", "fixture_swing", "fixture_sentiment"]
                     if c in self.comparison.columns]
        print(self.comparison[show_cols].to_string(index=False))

        # --- Upcoming fixtures per player ---
        print("\nUpcoming fixtures:")
        for _, player in self.comparison.iterrows():
            detail = player.get("upcoming_fixtures_detail", [])
            if detail:
                print(f"\n  {player['name']}:")
                for fx in detail[:5]:
                    ha = "🏠" if fx.get("is_home") else "✈️"
                    print(f"    GW{fx.get('gameweek','?')} {ha} {fx.get('opponent_name','?')} "
                          f"— {fx.get('difficulty','?')}")

        # --- Transfer candidates ---
        print(f"\n{'='*70}")
        print("Transfer candidates:")
        self.transfer_targets = self.analyzer.identify_transfer_candidates(self.comparison)
        if not self.transfer_targets.empty:
            print(self.transfer_targets.to_string(index=False))
        else:
            print("  ✓ No urgent transfers flagged.")

        # --- Export Excel ---
        excel_path = os.path.join(self.output_dir, "fpl_analysis.xlsx")
        ExcelExporter.export(self.comparison, self.transfer_targets, excel_path)

        # --- ML recommendations ---
        print(f"\n{'='*70}")
        print("Training ML model …")
        self.recommender.train(self.players_df)
        budget = (current_gw["value"] + current_gw["bank"]) / 10

        for pos_id, pos_name in self.POSITION_NAMES.items():
            print(f"\n  {pos_name}S:")
            recs = self.recommender.recommend(
                self.players_df, current_ids, pos_id, budget,
                n=3, current_gw=current_gw["gameweek"]
            )
            if not recs.empty:
                print(recs.to_string(index=False))
                self.all_recommendations[pos_name] = recs

        return self.all_recommendations

    def start_chat(self) -> None:
        """Interactive Q&A loop powered by ChatGPT."""
        if not self.chatgpt:
            print("\n[INFO] OpenAI key not set — chat unavailable.")
            return

        recs_text = "\n\n".join(
            f"{pos}:\n{df.to_string(index=False)}"
            for pos, df in self.all_recommendations.items()
        )
        context = self.chatgpt.build_context(
            self.comparison, self.transfer_targets, recs_text, self.players_df
        )

        print("\n" + "="*70)
        print("FPL Chat — type 'exit' to quit")
        print("="*70)
        first = True
        while True:
            msg = input("\nYou: ").strip()
            if msg.lower() in ("exit", "quit", "bye"):
                print("Good luck this GW!")
                break
            if not msg:
                continue
            reply = self.chatgpt.chat(msg, context if first else None, self.players_df)
            print(f"\nAssistant: {reply}")
            first = False

    def export_data(self) -> None:
        if self.team_data is not None:
            path = os.path.join(self.output_dir, "gameweek_history.csv")
            self.team_data.to_csv(path, index=False)
            print(f"  ✓ Gameweek history → {path}")


# ===========================================================================
# Entry point
# ===========================================================================

if __name__ == "__main__":
    # -----------------------------------------------------------------------
    # Configuration — edit these two values or set environment variables.
    # -----------------------------------------------------------------------
    TEAM_ID: int = int(os.getenv("FPL_TEAM_ID", "0"))
    OPENAI_KEY: str | None = os.getenv("OPENAI_API_KEY")   # None = skip chat features

    if TEAM_ID == 0:
        print("Set your FPL_TEAM_ID environment variable or edit TEAM_ID above.")
    else:
        assistant = FPLAssistant(TEAM_ID, openai_api_key=OPENAI_KEY)
        assistant.initialize()
        assistant.run_analysis()
        assistant.export_data()
        assistant.start_chat()  # no-op if OPENAI_KEY is not set
